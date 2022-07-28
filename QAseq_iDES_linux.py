import pandas as pd
import os
from pandas.core.frame import DataFrame
import openpyxl
import glob
import argparse

"""
Usage: get varinat information on target sites from mutation excels
Input: several excel files with one excel as reference:
Output: another sheet for each sample mutation excel
"""

def extract_mutation(input_path: str, sample_name: str):
    mutation_path = input_path + "/mutation"
    print("mutation files in: "+mutation_path)
    os.chdir(mutation_path)
    ref_path = input_path + "/reference/*"
    ref_file = glob.glob(ref_path)[0]
    print("the reference file is: "+ref_file)
    fileNames=[]
    sample_file = mutation_path + "/" + "*-" + sample_name + "*varlist.xlsx"
    for file in glob.glob(sample_file):
        fileNames.append(file)
    
    for working_file in fileNames:
        print("working on: "+working_file)
        xls_ref = pd.ExcelFile(ref_file)
        df_ref = pd.read_excel(xls_ref, 'final')
        df_ref.rename(columns={"Unnamed: 0":"index"}, inplace=True)
        xls_work = pd.ExcelFile(working_file)
        df_work = pd.read_excel(xls_work, 'raw')
        df_work.rename(columns={"Unnamed: 0":"index"}, inplace=True)
        MRD_real = []
        MRD_error = []
        
        
        # for normal results
        for _, row1 in df_ref.iterrows():
            for _, row2 in df_work.iterrows():
                if row1['chr'] == row2['chr'] and row1["start"] == row2["start"] and row1["stop"] == row2["stop"]:
                    if row1["var"] == row2["var"]:
                        MRD_real.append(row2)
                    elif row1["var"] != row2["var"]:
                        MRD_error.append(row2)
        
        '''
        # for cmm ref results
        for _, row1 in df_ref.iterrows():
            for _, row2 in df_work.iterrows():
                if row1['chr'] == row2['chr'] and row1["start"] == row2["start"]:
                    if row1["ref"] == row2["var"]:
                        MRD_real.append(row2)
                    elif row1["ref"] != row2["var"]:
                        MRD_error.append(row2)
        '''
        # Only export real MRD
        MRD_df = DataFrame(MRD_real)
        MRD_error_df = DataFrame(MRD_error)
        #MRD_df_all = pd.concat([MRD_df, MRD_error_df])
        
        # Use package openyxl to edit excel files
        wb = openpyxl.load_workbook(working_file)
        # Remove the existed "MRD" sheet
        if "MRD" in wb.sheetnames:
            ws = wb["MRD"]
            wb.remove(ws)
            wb.save(working_file)
        if "MRD_error" in wb.sheetnames:
            ws = wb["MRD_error"]
            wb.remove(ws)
            wb.save(working_file)
        with pd.ExcelWriter(working_file, mode="a", engine="openpyxl") as writer:
            MRD_df.to_excel(writer, sheet_name="MRD")
            MRD_error_df.to_excel(writer, sheet_name="MRD_error")
            """
            if "MRD_real" not in writer.sheets:
                MRD_df.to_excel(writer, sheet_name="MRD_real")
            if "MRD_error" not in writer.sheets:
                MRD_error_df.to_excel(writer, sheet_name="MRD_error")
            """

def create_FREQ(input_path: str, sample_name: str):
    # Specify working dir
    working_dir=input_path + "/freq_test"
    if not os.path.exists(working_dir):
                os.mkdir(working_dir)
    os.chdir(working_dir)

    # Specify reference and design table
    ref_path = input_path + "/reference/*"
    ref_file = glob.glob(ref_path)[0]
    design_file = input_path+"/design.xlsx"

    # Go through each file in mutation and depth file
    # Specify base name!
    mut_files = input_path + "/mutation/" + "*-" + sample_name + "*varlist.xlsx"
    fileNames=[]
    for file in glob.glob(mut_files):
        fileNames.append(file)


    for mut_file in fileNames:
        print(mut_file)
        depth_file = mut_file.replace("mutation", "datasummary").replace("_varlist", "_UMI_number")
        baseSuper = mut_file.split("/")[-1].split("_")[0]
        #depth_file = "C:\\Users\\admin\\documents\\MRD\\standardResult\\standard_3rdtrial\\datasummary\\79-NC\\1-79-NC12_UMI_number.xlsx"

        # store infos
        freq_dict = {}

        design_df = pd.read_excel(design_file)
        for index, row in design_df.iterrows():
            freq_dict[row.plex_id] = {}
            pos = row.ROI_start - row.start + 1
            freq_dict[row.plex_id]["POS"] = pos

        depth_df = pd.read_excel(depth_file)
        for index, row in depth_df.iterrows():
            freq_dict[row.amplicon_name]["DEPTH"] = row[2]

        ref_df = pd.read_excel(ref_file)
        for index, row in ref_df.iterrows():
            freq_dict[row.plex_id]["REF"] = row.ref

        mut_df_real = pd.read_excel(mut_file, "MRD")
        mut_df_fake = pd.read_excel(mut_file, "MRD_error")
        for index, row in mut_df_real.iterrows():
            if row["var"] == "A" and "A+" in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["A+"] += row.mol_count
            elif row["var"] == "A" and "A+" not in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["A+"] = row.mol_count
            if row["var"] == "C" and "C+" in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["C+"] += row.mol_count
            elif row["var"] == "C" and "C+" not in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["C+"] = row.mol_count
            if row["var"] == "T" and "T+" in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["T+"] += row.mol_count
            elif row["var"] == "T" and "T+" not in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["T+"] = row.mol_count
            if row["var"] == "G" and "G+" in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["G+"] += row.mol_count
            elif row["var"] == "G" and "G+" not in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["G+"] = row.mol_count

        for index, row in mut_df_fake.iterrows():
            if row["var"] == "A" and "A+" in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["A+"] += row.mol_count
            elif row["var"] == "A" and "A+" not in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["A+"] = row.mol_count
            if row["var"] == "C" and "C+" in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["C+"] += row.mol_count
            elif row["var"] == "C" and "C+" not in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["C+"] = row.mol_count
            if row["var"] == "T" and "T+" in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["T+"] += row.mol_count
            elif row["var"] == "T" and "T+" not in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["T+"] = row.mol_count
            if row["var"] == "G" and "G+" in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["G+"] += row.mol_count
            elif row["var"] == "G" and "G+" not in freq_dict[row.plex_id]:
                freq_dict[row.plex_id]["G+"] = row.mol_count

        for key, value in freq_dict.items():
            if "A+" not in value:
                value["A+"] = 0
            if "C+" not in value:
                value["C+"] = 0
            if "T+" not in value:
                value["T+"] = 0
            if "G+" not in value:
                value["G+"] = 0
            value["A-"] = 0
            value["C-"] = 0
            value["T-"] = 0
            value["G-"] = 0
            value["R-"] = 0
            value["R+"] = value["DEPTH"] - value["A+"] - value["C+"] - value["T+"] - value["G+"]

        freq_df = pd.DataFrame(freq_dict).T
        freq_df.reset_index(inplace=True)
        freq_df = freq_df.rename(columns = {'index':'CHR'})
        order = ["CHR", "POS", "DEPTH", "REF", "R+", "R-", "A+", "A-", "C+", "C-", "T+", "T-", "G+", "G-"]
        freq_df = freq_df[order]

        outputName = baseSuper+"_UMI_sorted.freq.allreads.Q30.txt"
        freq_df.to_csv(outputName, sep = "\t", index = False)



if __name__=='__main__':
        parser = argparse.ArgumentParser(description='the unmapped reads blat with hg38')
        parser.add_argument('-i', '--inputpath', type=str, help='the path of lane')
        parser.add_argument('-s', '--samplename', type=str, help='the sample ID name')
        parser.add_argument('-t', '--Thread1', type=int, help='Thread count')
        args = parser.parse_args()
        Inputpath = args.inputpath
        Samplename = args.samplename
        Thread1=args.Thread1
        if not Thread1:
                Thread1=5
        # Run modify mutation table
        #extract_mutation(Inputpath,Samplename)
        # Get UMI-level FREQ files
        create_FREQ(Inputpath,Samplename)

