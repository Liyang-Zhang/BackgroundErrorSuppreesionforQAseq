# -*- coding: utf-8 -*-
"""
Created on Fri Mar 18 12:59:24 2022

@author: Liyang Zhang
Usage: get varinat information on target sites from mutation excels
Input: several excel files with one excel as reference:
Output: another sheet for each sample mutation excel
"""

import pandas as pd
import os
from pandas.core.frame import DataFrame
import openpyxl
import glob

working_dir="C:\\Users\\admin\\documents\\MRD\\standardResult\\standard_3rdtrial\\mutation\\Q"
os.chdir(working_dir)
'''
reference for ABC: C:\\Users\\admin\\documents\\MRD\\standardResult\\new_design_3_21\\mutation\\H-1-reference.xlsx
reference for KLM: C:\\Users\\admin\\documents\\MRD\\standardResult\\standard_3rdtrial\\mutation\\K\\254676_KLM_reference.xlsx
reference for DEF: C:\\Users\\admin\\documents\\MRD\\standardResult\\standard_3rdtrial\\mutation\\F\\254674_DEF_reference.xlsx
reference for NPQ: C:\\Users\\admin\\documents\\MRD\\standardResult\\standard_3rdtrial\\mutation\\N\\254677_NPQ_reference.xlsx
reference for RST: C:\\Users\\admin\\documents\\MRD\\standardResult\\standard_3rdtrial\\mutation\\R\\254678_RST_reference.xlsx
reference for GHJ: C:\\Users\\admin\\documents\\MRD\\standardResult\\standard_3rdtrial\\mutation\\H\\254675_GHJ_reference.xlsx
'''
ref_file="C:\\Users\\admin\\documents\\MRD\\standardResult\\standard_3rdtrial\\mutation\\N\\254677_NPQ_reference.xlsx"
fileNames=[]
for file in glob.glob("*-Q*.xlsx"):
    fileNames.append(file)

for working_file in fileNames:
    # working_file="1-C1_varlist.xlsx"
    
    xls_ref = pd.ExcelFile(ref_file)
    df_ref = pd.read_excel(xls_ref, 'final')
    df_ref.rename(columns={"Unnamed: 0":"index"}, inplace=True)
    xls_work = pd.ExcelFile(working_file)
    df_work = pd.read_excel(xls_work, 'raw')
    df_work.rename(columns={"Unnamed: 0":"index"}, inplace=True)
    MRD_real = []
    MRD_error = []
    
    
    # for normal results
    for _, row1 in df_ref.iterrows():
        for _, row2 in df_work.iterrows():
            if row1['chr'] == row2['chr'] and row1["start"] == row2["start"] and row1["stop"] == row2["stop"]:
                if row1["var"] == row2["var"]:
                    MRD_real.append(row2)
                elif row1["var"] != row2["var"]:
                    MRD_error.append(row2)
    
    '''
    # for cmm ref results
    for _, row1 in df_ref.iterrows():
        for _, row2 in df_work.iterrows():
            if row1['chr'] == row2['chr'] and row1["start"] == row2["start"]:
                if row1["ref"] == row2["var"]:
                    MRD_real.append(row2)
                elif row1["ref"] != row2["var"]:
                    MRD_error.append(row2)
    '''
    # Only export real MRD
    MRD_df = DataFrame(MRD_real)
    MRD_error_df = DataFrame(MRD_error)
    #MRD_df_all = pd.concat([MRD_df, MRD_error_df])
    
    # Use package openyxl to edit excel files
    wb = openpyxl.load_workbook(working_file)
    # Remove the existed "MRD" sheet
    if "MRD" in wb.sheetnames:
        ws = wb["MRD"]
        wb.remove(ws)
        wb.save(working_file)
    if "MRD_error" in wb.sheetnames:
        ws = wb["MRD_error"]
        wb.remove(ws)
        wb.save(working_file)
    with pd.ExcelWriter(working_file, mode="a", engine="openpyxl") as writer:
        MRD_df.to_excel(writer, sheet_name="MRD")
        MRD_error_df.to_excel(writer, sheet_name="MRD_error")
        """
        if "MRD_real" not in writer.sheets:
            MRD_df.to_excel(writer, sheet_name="MRD_real")
        if "MRD_error" not in writer.sheets:
            MRD_error_df.to_excel(writer, sheet_name="MRD_error")
        """
